from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.schemas.results_export import ExportHistoryEntryModel


NAVY = "0F2747"
BLUE = "1E5AA8"
SKY = "D9E8FB"
GOLD = "C89A48"
INK = "1C2430"
SOFT = "F6F8FC"
WHITE = "FFFFFF"
GRID = "D7DFEB"
MINT = "E9F5EE"
ROSE = "FCECEF"
PALE = "F4F7FB"

PROJECT_ROOT = Path(__file__).resolve().parents[2]
LOGO_PATH = PROJECT_ROOT / "app" / "web" / "static" / "assets" / "mad-formula-team-logo.png"

THIN_BORDER = Border(
    left=Side(style="thin", color=GRID),
    right=Side(style="thin", color=GRID),
    top=Side(style="thin", color=GRID),
    bottom=Side(style="thin", color=GRID),
)


def _group_entries(entries: list[ExportHistoryEntryModel]) -> dict[int, list[ExportHistoryEntryModel]]:
    grouped: dict[int, list[ExportHistoryEntryModel]] = {}
    for entry in sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True):
        grouped.setdefault(entry.summary.visible_layers, []).append(entry)
    return dict(sorted(grouped.items(), key=lambda item: item[0]))


def _format_theta(theta: object) -> str:
    value = float(theta)
    if abs(value) < 1e-12:
        return "0 deg"
    if abs(value).is_integer():
        value_text = f"{int(abs(value))}"
    else:
        value_text = f"{abs(value):.3f}".rstrip("0").rstrip(".")
    if value > 0:
        return f"+-{value_text} deg"
    return f"-{value_text} deg"


def _build_laminate_text(entry: ExportHistoryEntryModel) -> str:
    layers = entry.form_state.get("layers", [])
    orientations = []
    for layer in layers:
        if isinstance(layer, dict):
            orientations.append(_format_theta(layer.get("theta_deg", 0)))
    sequence = ", ".join(orientations)
    suffix = " S" if entry.summary.is_symmetric else ""
    return f"[{sequence}]{suffix}"


def _build_cf_type(entry: ExportHistoryEntryModel) -> str:
    seen: list[str] = []
    for layer in entry.form_state.get("layers", []):
        if isinstance(layer, dict):
            material_id = str(layer.get("material_id", ""))
            if material_id and material_id not in seen:
                seen.append(material_id)
    return ", ".join(seen)


def _core_thickness(entry: ExportHistoryEntryModel) -> float:
    multiplier = 2 if entry.summary.is_symmetric else 1
    return max(entry.summary.total_thickness_mm - multiplier * entry.summary.fiber_thickness_mm, 0.0)


def _layer_count_label(layer_count: int) -> str:
    return f"{layer_count} capas"


def _entry_name(entry: ExportHistoryEntryModel, layer_count: int, index: int) -> str:
    timestamp = ""
    if entry.saved_at:
        try:
            timestamp = datetime.fromisoformat(entry.saved_at).strftime("%m%d_%H%M")
        except ValueError:
            timestamp = ""
    suffix = f"_{timestamp}" if timestamp else ""
    return f"P{layer_count}_{index:02d}{suffix}"


def _set_sheet_title_block(worksheet, title: str, subtitle: str) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:H2")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(name="Montserrat", size=18, bold=True, color=WHITE)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor=NAVY)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A3:H4")
    worksheet["A3"] = subtitle
    worksheet["A3"].font = Font(name="Open Sans", size=10, italic=True, color=INK)
    worksheet["A3"].fill = PatternFill(fill_type="solid", fgColor=SKY)
    worksheet["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.merge_cells("I1:J4")
    worksheet["I1"] = ""
    worksheet["I1"].fill = PatternFill(fill_type="solid", fgColor=WHITE)
    worksheet["I1"].border = THIN_BORDER

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 34
    worksheet.row_dimensions[4].height = 34


def _add_logo(worksheet, cell: str = "I1", *, width: int = 120, height: int = 120) -> None:
    if not LOGO_PATH.exists():
        return
    logo = Image(str(LOGO_PATH))
    logo.width = width
    logo.height = height
    worksheet.add_image(logo, cell)


def _style_kpi_box(worksheet, cell_range: str, title: str, value: str, *, fill_color: str, value_color: str = INK) -> None:
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    worksheet[start_cell] = f"{title}\n{value}"
    worksheet[start_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet[start_cell].fill = PatternFill(fill_type="solid", fgColor=fill_color)
    worksheet[start_cell].font = Font(name="Montserrat", bold=True, size=13, color=value_color)
    worksheet[start_cell].border = THIN_BORDER


def _style_info_panel(worksheet, cell_range: str, title: str, body: str) -> None:
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    worksheet[start_cell] = f"{title}\n\n{body}"
    worksheet[start_cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    worksheet[start_cell].fill = PatternFill(fill_type="solid", fgColor=WHITE)
    worksheet[start_cell].font = Font(name="Open Sans", size=10, color=INK)
    worksheet[start_cell].border = THIN_BORDER


def _configure_chart_axes(chart, x_title: str, y_title: str) -> None:
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.legend = None
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.numFmt = "0.000"
    if getattr(chart.y_axis, "majorGridlines", None) is not None:
        chart.y_axis.majorGridlines = None


def _add_chart_value_labels(chart) -> None:
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = False
    chart.dLbls.showLegendKey = False
    chart.dLbls.showSerName = False
    chart.visible_cells_only = False


def _set_string_categories(chart, worksheet_title: str, column: int, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    column_letter = get_column_letter(column)
    reference = f"'{worksheet_title}'!${column_letter}${start_row}:${column_letter}${end_row}"
    category_source = AxDataSource(strRef=StrRef(f=reference))
    for series in chart.ser:
        series.cat = category_source


def _style_header_row(worksheet, row: int, max_col: int) -> None:
    worksheet.row_dimensions[row].height = 24
    for column in range(1, max_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.font = Font(name="Montserrat", bold=True, color=WHITE)
        cell.fill = PatternFill(fill_type="solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_region(worksheet, start_row: int, end_row: int, max_col: int) -> None:
    for row in range(start_row, end_row + 1):
        fill = PatternFill(fill_type="solid", fgColor=WHITE if row % 2 else SOFT)
        max_length = 0
        for column in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.font = Font(name="Open Sans", color=INK)
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        worksheet.row_dimensions[row].height = max(22, min(54, 18 + (max_length // 28) * 10))


def _auto_width(worksheet, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _add_table(worksheet, start_row: int, end_row: int, end_col: int, table_name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _write_summary_sheet(workbook: Workbook, entries: list[ExportHistoryEntryModel]) -> None:
    worksheet = workbook.create_sheet("Resumen")
    worksheet.sheet_properties.tabColor = BLUE
    _set_sheet_title_block(
        worksheet,
        "Resultados de sesión",
        "Comparativa global",
    )

    helper_start_col = 14
    worksheet.column_dimensions[get_column_letter(helper_start_col)].hidden = True
    worksheet.column_dimensions[get_column_letter(helper_start_col + 1)].hidden = True
    worksheet.column_dimensions[get_column_letter(helper_start_col + 2)].hidden = True
    worksheet.cell(row=6, column=helper_start_col).value = "Configuración"
    worksheet.cell(row=6, column=helper_start_col + 1).value = "Elastic gradient theory"
    worksheet.cell(row=6, column=helper_start_col + 2).value = "EI theory"

    headers = [
        "Nombre",
        "Capas visibles",
        "Laminate",
        "CF type",
        "Core",
        "Espesor fibra (mm)",
        "Espesor total (mm)",
        "Elastic gradient theory",
        "EI theory",
        "Fecha",
    ]
    start_row = 29
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=column).value = header
    _style_header_row(worksheet, start_row, len(headers))

    ordered_entries = sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True)
    for row_index, entry in enumerate(ordered_entries, start=start_row + 1):
        helper_row = row_index - start_row + 6
        worksheet.cell(row=helper_row, column=helper_start_col).value = _entry_name(entry, layer_count=entry.summary.visible_layers, index=row_index - start_row)
        worksheet.cell(row=helper_row, column=helper_start_col + 1).value = round(entry.summary.elastic_gradient_theory, 3)
        worksheet.cell(row=helper_row, column=helper_start_col + 2).value = round(entry.summary.ei_theory, 3)
        layer_count = entry.summary.visible_layers
        worksheet.cell(row=row_index, column=1).value = _entry_name(entry, layer_count, row_index - start_row)
        worksheet.cell(row=row_index, column=2).value = layer_count
        worksheet.cell(row=row_index, column=3).value = _build_laminate_text(entry)
        worksheet.cell(row=row_index, column=4).value = _build_cf_type(entry)
        worksheet.cell(row=row_index, column=5).value = entry.summary.core_material_id
        worksheet.cell(row=row_index, column=6).value = round(entry.summary.fiber_thickness_mm, 3)
        worksheet.cell(row=row_index, column=7).value = round(entry.summary.total_thickness_mm, 3)
        worksheet.cell(row=row_index, column=8).value = round(entry.summary.elastic_gradient_theory, 3)
        worksheet.cell(row=row_index, column=9).value = round(entry.summary.ei_theory, 3)
        worksheet.cell(row=row_index, column=10).value = entry.saved_at or ""

    end_row = start_row + len(ordered_entries)
    _style_data_region(worksheet, start_row + 1, end_row, len(headers))
    _add_table(worksheet, start_row, end_row, len(headers), "ResumenResultados")
    _auto_width(
        worksheet,
        {
            1: 18,
            2: 14,
            3: 42,
            4: 18,
            5: 18,
            6: 18,
            7: 18,
            8: 24,
            9: 16,
            10: 22,
        },
    )
    worksheet.freeze_panes = f"A{start_row + 1}"
    worksheet.row_dimensions[start_row].height = 26

    if ordered_entries:
        bar_chart = BarChart()
        bar_chart.title = "Elastic gradient theory por configuración"
        _configure_chart_axes(bar_chart, "Configuraciones", "Elastic gradient theory")
        _add_chart_value_labels(bar_chart)
        data = Reference(worksheet, min_col=helper_start_col + 1, min_row=6, max_row=6 + len(ordered_entries))
        labels = Reference(worksheet, min_col=helper_start_col, min_row=7, max_row=6 + len(ordered_entries))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(labels)
        _set_string_categories(bar_chart, worksheet.title, helper_start_col, 7, 6 + len(ordered_entries))
        bar_chart.height = 10
        bar_chart.width = 20
        bar_chart.style = 10
        worksheet.add_chart(bar_chart, "A6")


def _write_cover_sheet(workbook: Workbook, entries: list[ExportHistoryEntryModel]) -> None:
    worksheet = workbook.active
    worksheet.title = "Portada"
    worksheet.sheet_properties.tabColor = NAVY
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells("A1:H2")
    worksheet["A1"] = "MAD Formula Team"
    worksheet["A1"].font = Font(name="Montserrat", size=22, bold=True, color=WHITE)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor=NAVY)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A3:H4")
    worksheet["A3"] = "Resultados de laminados"
    worksheet["A3"].font = Font(name="Montserrat", size=18, bold=True, color=INK)
    worksheet["A3"].fill = PatternFill(fill_type="solid", fgColor=SKY)
    worksheet["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.merge_cells("I1:J4")
    worksheet["I1"] = ""
    worksheet["I1"].fill = PatternFill(fill_type="solid", fgColor=WHITE)
    worksheet["I1"].border = THIN_BORDER
    _add_logo(worksheet, "I1", width=95, height=95)

    ordered_entries = sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True)
    grouped = _group_entries(entries)
    best_entry = max(entries, key=lambda item: item.summary.elastic_gradient_theory)
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    _style_kpi_box(worksheet, "A6:C8", "Configuraciones", str(len(entries)), fill_color=SOFT)
    _style_kpi_box(
        worksheet,
        "D6:F8",
        "Mejor Elastic gradient theory",
        f"{best_entry.summary.elastic_gradient_theory:.3f}",
        fill_color=MINT,
    )
    _style_kpi_box(worksheet, "G6:H8", "Grupos de capas", str(len(grouped)), fill_color=SOFT)
    _style_kpi_box(worksheet, "I6:J8", "Exportado", exported_at, fill_color=ROSE)

    _style_info_panel(
        worksheet,
        "A10:E16",
        "Mejor configuración de la sesión",
        (
            f"Laminate: {_build_laminate_text(best_entry)}\n"
            f"CF type: {_build_cf_type(best_entry)}\n"
            f"Core: {best_entry.summary.core_material_id}\n"
            f"EI theory: {best_entry.summary.ei_theory:.3f}\n"
            f"Elastic gradient theory: {best_entry.summary.elastic_gradient_theory:.3f}"
        ),
    )
    _style_info_panel(
        worksheet,
        "F10:J16",
        "Qué incluye este libro",
        (
            "Portada ejecutiva con resumen de sesión.\n"
            "Hoja Resumen con comparativa global y gráficos.\n"
            "Una hoja por número de capas presente en la sesión.\n"
            "Hoja Metadatos con trazabilidad completa para análisis y exportaciones futuras."
        ),
    )
    worksheet["A10"].fill = PatternFill(fill_type="solid", fgColor=PALE)
    worksheet["F10"].fill = PatternFill(fill_type="solid", fgColor=PALE)
    for row in range(10, 17):
        worksheet.row_dimensions[row].height = 26

    start_row = 19
    headers = ["Grupo", "Configuraciones", "Max Elastic gradient theory", "Max EI theory"]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=column).value = header
    _style_header_row(worksheet, start_row, len(headers))

    group_items = sorted(grouped.items(), key=lambda item: item[0])
    for row_index, (layer_count, group_entries) in enumerate(group_items, start=start_row + 1):
        worksheet.cell(row=row_index, column=1).value = _layer_count_label(layer_count)
        worksheet.cell(row=row_index, column=2).value = len(group_entries)
        worksheet.cell(row=row_index, column=3).value = round(max(item.summary.elastic_gradient_theory for item in group_entries), 3)
        worksheet.cell(row=row_index, column=4).value = round(max(item.summary.ei_theory for item in group_entries), 3)
    end_row = start_row + len(group_items)
    _style_data_region(worksheet, start_row + 1, end_row, len(headers))
    _add_table(worksheet, start_row, end_row, len(headers), "ResumenGrupos")

    _auto_width(
        worksheet,
        {
            1: 18,
            2: 18,
            3: 28,
            4: 18,
            5: 16,
            6: 16,
            7: 16,
            8: 16,
            9: 16,
            10: 16,
        },
    )

    if group_items:
        chart = BarChart()
        chart.title = "Mejor Elastic gradient theory por grupo"
        _configure_chart_axes(chart, "Numero de capas", "Elastic gradient theory")
        _add_chart_value_labels(chart)
        data = Reference(worksheet, min_col=3, min_row=start_row, max_row=end_row)
        labels = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        _set_string_categories(chart, worksheet.title, 1, start_row + 1, end_row)
        chart.height = 10
        chart.width = 18
        chart.style = 10
        worksheet.add_chart(chart, "F19")


def _write_group_sheet(workbook: Workbook, layer_count: int, entries: list[ExportHistoryEntryModel]) -> None:
    worksheet = workbook.create_sheet(_layer_count_label(layer_count))
    worksheet.sheet_properties.tabColor = GOLD
    _set_sheet_title_block(
        worksheet,
        f"Resultados para {_layer_count_label(layer_count)}",
        "Ordenados de mayor a menor por EI theory para facilitar comparación y selección",
    )

    _add_logo(worksheet, "L1", width=78, height=78)

    ordered_entries = sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True)
    helper_start_col = 15
    worksheet.column_dimensions[get_column_letter(helper_start_col)].hidden = True
    worksheet.column_dimensions[get_column_letter(helper_start_col + 1)].hidden = True
    worksheet.cell(row=6, column=helper_start_col).value = "Configuración"
    worksheet.cell(row=6, column=helper_start_col + 1).value = "Elastic gradient theory"
    data_start_row = 29

    if ordered_entries:
        bar_chart = BarChart()
        bar_chart.title = f"Elastic gradient theory - {_layer_count_label(layer_count)}"
        _configure_chart_axes(bar_chart, "Configuraciones", "Elastic gradient theory")
        _add_chart_value_labels(bar_chart)
        bar_chart.height = 10
        bar_chart.width = 20
        bar_chart.style = 11

        for row_index, entry in enumerate(ordered_entries, start=7):
            worksheet.cell(row=row_index, column=helper_start_col).value = _entry_name(entry, layer_count, row_index - 6)
            worksheet.cell(row=row_index, column=helper_start_col + 1).value = round(entry.summary.elastic_gradient_theory, 3)

        data = Reference(worksheet, min_col=helper_start_col + 1, min_row=6, max_row=6 + len(ordered_entries))
        labels = Reference(worksheet, min_col=helper_start_col, min_row=7, max_row=6 + len(ordered_entries))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(labels)
        _set_string_categories(bar_chart, worksheet.title, helper_start_col, 7, 6 + len(ordered_entries))
        worksheet.add_chart(bar_chart, "A6")

    headers = [
        "Nombre",
        "Tipo",
        "CF type",
        "Laminate",
        "Core material",
        "Core thickness (mm)",
        "Espesor total (mm)",
        "Espesor de fibra (mm)",
        "Elastic gradient theory",
        "EI theory",
        "Simetría",
        "Fecha",
    ]
    start_row = data_start_row
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=column).value = header
    _style_header_row(worksheet, start_row, len(headers))

    for row_index, entry in enumerate(ordered_entries, start=start_row + 1):
        worksheet.cell(row=row_index, column=1).value = _entry_name(entry, layer_count, row_index - start_row)
        worksheet.cell(row=row_index, column=2).value = "3PBT"
        worksheet.cell(row=row_index, column=3).value = _build_cf_type(entry)
        worksheet.cell(row=row_index, column=4).value = _build_laminate_text(entry)
        worksheet.cell(row=row_index, column=5).value = entry.summary.core_material_id
        worksheet.cell(row=row_index, column=6).value = round(_core_thickness(entry), 3)
        worksheet.cell(row=row_index, column=7).value = round(entry.summary.total_thickness_mm, 3)
        worksheet.cell(row=row_index, column=8).value = round(entry.summary.fiber_thickness_mm, 3)
        worksheet.cell(row=row_index, column=9).value = round(entry.summary.elastic_gradient_theory, 3)
        worksheet.cell(row=row_index, column=10).value = round(entry.summary.ei_theory, 3)
        worksheet.cell(row=row_index, column=11).value = "Si" if entry.summary.is_symmetric else "No"
        worksheet.cell(row=row_index, column=12).value = entry.saved_at or ""

    end_row = start_row + len(ordered_entries)
    _style_data_region(worksheet, start_row + 1, end_row, len(headers))
    _add_table(worksheet, start_row, end_row, len(headers), f"Grupo{layer_count}Capas")
    _auto_width(
        worksheet,
        {
            1: 18,
            2: 12,
            3: 18,
            4: 42,
            5: 18,
            6: 18,
            7: 16,
            8: 18,
            9: 24,
            10: 16,
            11: 12,
            12: 22,
        },
    )
    worksheet.freeze_panes = f"A{start_row + 1}"
    worksheet.row_dimensions[start_row].height = 26


def _write_metadata_sheet(workbook: Workbook, entries: list[ExportHistoryEntryModel]) -> None:
    worksheet = workbook.create_sheet("Metadatos")
    worksheet.sheet_properties.tabColor = NAVY
    worksheet.sheet_state = "hidden"
    _set_sheet_title_block(
        worksheet,
        "Metadatos completos",
        "Hoja de trazabilidad con configuración, resumen y resultado completo serializado",
    )

    headers = [
        "saved_at",
        "signature",
        "visible_layers",
        "is_symmetric",
        "core_material_id",
        "fiber_thickness_mm",
        "total_thickness_mm",
        "elastic_gradient_theory",
        "ei_theory",
        "laminate_text",
        "cf_type",
        "layers_json",
        "form_state_json",
        "result_data_json",
    ]
    start_row = 6
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=column).value = header
    _style_header_row(worksheet, start_row, len(headers))

    ordered_entries = sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True)
    for row_index, entry in enumerate(ordered_entries, start=start_row + 1):
        worksheet.cell(row=row_index, column=1).value = entry.saved_at or ""
        worksheet.cell(row=row_index, column=2).value = entry.signature
        worksheet.cell(row=row_index, column=3).value = entry.summary.visible_layers
        worksheet.cell(row=row_index, column=4).value = "Si" if entry.summary.is_symmetric else "No"
        worksheet.cell(row=row_index, column=5).value = entry.summary.core_material_id
        worksheet.cell(row=row_index, column=6).value = round(entry.summary.fiber_thickness_mm, 3)
        worksheet.cell(row=row_index, column=7).value = round(entry.summary.total_thickness_mm, 3)
        worksheet.cell(row=row_index, column=8).value = round(entry.summary.elastic_gradient_theory, 3)
        worksheet.cell(row=row_index, column=9).value = round(entry.summary.ei_theory, 3)
        worksheet.cell(row=row_index, column=10).value = _build_laminate_text(entry)
        worksheet.cell(row=row_index, column=11).value = _build_cf_type(entry)
        worksheet.cell(row=row_index, column=12).value = json.dumps(entry.form_state.get("layers", []), ensure_ascii=False)
        worksheet.cell(row=row_index, column=13).value = json.dumps(entry.form_state, ensure_ascii=False)
        worksheet.cell(row=row_index, column=14).value = json.dumps(entry.result_data, ensure_ascii=False)

    end_row = start_row + len(ordered_entries)
    _style_data_region(worksheet, start_row + 1, end_row, len(headers))
    _add_table(worksheet, start_row, end_row, len(headers), "MetadatosResultados")
    _auto_width(
        worksheet,
        {
            1: 22,
            2: 26,
            3: 14,
            4: 12,
            5: 18,
            6: 18,
            7: 18,
            8: 24,
            9: 16,
            10: 42,
            11: 18,
            12: 28,
            13: 34,
            14: 46,
        },
    )
    worksheet.freeze_panes = "A7"


def build_results_export_workbook(entries: list[ExportHistoryEntryModel]) -> bytes:
    workbook = Workbook()

    ordered_entries = sorted(entries, key=lambda item: item.summary.ei_theory, reverse=True)
    grouped = _group_entries(ordered_entries)

    _write_cover_sheet(workbook, ordered_entries)
    _write_summary_sheet(workbook, ordered_entries)
    for layer_count, group_entries in grouped.items():
        _write_group_sheet(workbook, layer_count, group_entries)
    _write_metadata_sheet(workbook, ordered_entries)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_export_filename() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"resultados_sesion_{timestamp}.xlsx"
    worksheet.row_dimensions[3].height = 38
    worksheet.row_dimensions[4].height = 38
