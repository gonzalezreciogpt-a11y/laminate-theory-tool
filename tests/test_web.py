from fastapi.testclient import TestClient
from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.main import app


client = TestClient(app)


def test_healthcheck_endpoint() -> None:
    response = client.get("/healthz")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-frame-options"] == "DENY"
    assert "frame-ancestors 'none'" in response.headers["content-security-policy"]


def test_home_page_renders() -> None:
    response = client.get("/")
    assert response.status_code == 200
    assert "Calculadora de laminados." in response.text
    assert "Materiales disponibles" in response.text
    assert "Servicio público: los cálculos se ejecutan en el servidor" in response.text
    assert "Laminado simétrico" not in response.text
    assert 'name="is_symmetric"' not in response.text
    assert "Espesor del core (mm)" in response.text
    assert "material-accordion" in response.text
    assert "Ingredientes disponibles" not in response.text
    assert "Todo el flujo principal queda a la vista" not in response.text
    assert "Uso guiado" not in response.text


def test_api_calculate_returns_reference_shape() -> None:
    payload = {
        "layers": [
            {"material_id": "RC416T", "theta_deg": 45.0},
            {"material_id": "UD", "theta_deg": 0.0},
            {"material_id": "RC416T", "theta_deg": 90.0},
            {"material_id": "Dummy", "theta_deg": 0.0},
        ],
        "is_symmetric": True,
        "core_material_id": "Honeycomb",
        "compatibility_mode": "legacy",
    }
    response = client.post("/api/calculate", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert round(data["equivalent_properties"]["e11_gpa"], 6) == 66.152367
    assert round(data["three_point_bending"]["ei_theory"], 6) == 4724.34448


def test_form_post_matches_matlab_g12g_when_legacy_dummy_is_enabled() -> None:
    response = client.post(
        "/",
        data={
            "material_id": ["RC416T", "UD", "RC416T"],
            "theta_deg": ["45", "0", "90"],
            "is_symmetric": "on",
            "core_material_id": "Honeycomb",
            "insert_dummy_layer_for_odd_compatibility": "on",
            "elastic_gradient": "2649",
            "rigidez_rig": "14871",
            "span_m": "0.4",
            "span_mm": "400",
            "width_m": "0.275",
            "width_mm": "275",
        },
    )
    assert response.status_code == 200
    assert "33.076 GPa" in response.text
    assert "Bloque experimental" not in response.text
    assert "Propiedades del laminado" in response.text
    assert "Apilado final" in response.text
    assert "22.320 mm" in response.text
    assert "Traza CLT del sandwich visible" in response.text
    assert "Traza legado MATLAB" in response.text
    assert response.text.count('class="laminate-segment') >= 7


def test_materials_library_page_renders() -> None:
    response = client.get("/materials-library")
    assert response.status_code == 200
    assert "Biblioteca de materiales." in response.text
    assert "Catálogo del proyecto" in response.text
    assert "MD Balsa Wood" in response.text


def test_tutorial_page_renders() -> None:
    response = client.get("/tutorial")
    assert response.status_code == 200
    assert "Guía de uso." in response.text
    assert "Cómo trabajar con la app" in response.text


def test_results_page_renders() -> None:
    response = client.get("/results")
    assert response.status_code == 200
    assert "Resultados guardados." in response.text
    assert "Historial de resultados" in response.text
    assert "hero-actions hero-actions-wide" in response.text
    assert "result-group-details" in response.text


def test_compare_page_renders() -> None:
    response = client.get("/compare")
    assert response.status_code == 200
    assert "Comparador de resultados." in response.text
    assert "Comparativa activa" in response.text


def test_shuffle_page_renders() -> None:
    response = client.get("/shuffle")
    assert response.status_code == 200
    assert "Shuffle de laminados." in response.text
    assert "Shuffle guiado" in response.text
    assert "Capas a variar" in response.text
    assert "Todas las capas" in response.text
    assert 'id="shuffle-angle-checklist"' in response.text
    assert "shuffle-angles" not in response.text


def test_api_calculate_accepts_custom_materials() -> None:
    payload = {
        "layers": [
            {"material_id": "CUSTOM_SKIN", "theta_deg": 45.0},
            {"material_id": "RC416T", "theta_deg": 45.0},
        ],
        "is_symmetric": False,
        "core_material_id": "Honeycomb",
        "compatibility_mode": "legacy",
        "custom_materials": [
            {
                "id": "CUSTOM_SKIN",
                "name": "Custom Skin",
                "e1_pa": 70000000000.0,
                "e2_pa": 5000000000.0,
                "g12_pa": 3000000000.0,
                "poisson_input": 0.28,
                "strength_x": 0.0,
                "strength_x_compression": 0.0,
                "strength_y": 0.0,
                "strength_y_compression": 0.0,
                "strength_s": 0.0,
                "thickness_mm": 0.25,
                "user_selectable": True,
                "notes": "Temporary browser-local material",
                "fiber_family": "twill",
            }
        ],
    }
    response = client.post("/api/calculate", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert "CUSTOM_SKIN" in data["materials_catalog_used"]


def test_api_batch_calculate_returns_entries() -> None:
    payload = {
        "entries": [
            {
                "label": "Variante 1",
                "request": {
                    "layers": [
                        {"material_id": "RC416T", "theta_deg": 45.0},
                        {"material_id": "UD", "theta_deg": 0.0},
                        {"material_id": "RC416T", "theta_deg": 90.0},
                    ],
                    "is_symmetric": True,
                    "core_material_id": "Honeycomb",
                    "insert_dummy_layer_for_odd_compatibility": True,
                    "compatibility_mode": "legacy",
                    "custom_materials": [],
                },
            },
            {
                "label": "Variante 2",
                "request": {
                    "layers": [
                        {"material_id": "RC416T", "theta_deg": 45.0},
                        {"material_id": "UD", "theta_deg": 0.0},
                        {"material_id": "RC416T", "theta_deg": 90.0},
                    ],
                    "is_symmetric": True,
                    "core_material_id": "Honeycomb",
                    "insert_dummy_layer_for_odd_compatibility": True,
                    "compatibility_mode": "legacy",
                    "custom_materials": [],
                },
            },
        ]
    }

    response = client.post("/api/batch-calculate", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert len(data["entries"]) == 2
    assert data["entries"][0]["label"] == "Variante 1"
    assert round(data["entries"][0]["result"]["three_point_bending"]["elastic_gradient_theory"], 3) > 0


def test_api_export_results_returns_dynamic_workbook() -> None:
    payload = {
        "entries": [
            {
                "signature": "test-signature",
                "saved_at": "2026-04-18T12:00:00",
                "form_state": {
                    "layers": [
                        {"material_id": "RC416T", "theta_deg": 45.0},
                        {"material_id": "UD", "theta_deg": 0.0},
                        {"material_id": "RC416T", "theta_deg": 90.0},
                    ],
                    "is_symmetric": True,
                    "core_material_id": "Honeycomb",
                    "insert_dummy_layer_for_odd_compatibility": True,
                    "compatibility_mode": "legacy",
                    "custom_materials": [],
                },
                "summary": {
                    "elastic_gradient_theory": 3543.25836,
                    "ei_theory": 4724.34448,
                    "fiber_thickness_mm": 1.16,
                    "total_thickness_mm": 22.32,
                    "core_material_id": "Honeycomb",
                    "is_symmetric": True,
                    "visible_layers": 3,
                },
                "result_data": {
                    "equivalent_properties": {
                        "e11_gpa": 66.152367,
                    }
                },
            }
        ]
    }

    response = client.post("/api/export-results", json=payload)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Portada", "Resumen", "3 capas", "Metadatos"]
    worksheet = workbook["3 capas"]
    assert worksheet["E30"].value == "Honeycomb"
    assert worksheet["G30"].value == 22.32
    assert worksheet["H30"].value == 1.16
    assert worksheet["I30"].value == 3543.258
    assert workbook["Portada"]["A1"].value == "MAD Formula Team"
    assert workbook["Portada"]["A3"].value == "Resultados de laminados"
    assert workbook["Resumen"]["A3"].value == "Comparativa global"
    assert workbook["Metadatos"].sheet_state == "hidden"
    assert "Metadatos" in workbook.sheetnames

    with ZipFile(BytesIO(response.content)) as archive:
        chart_files = [name for name in archive.namelist() if name.startswith("xl/charts/chart")]
        assert len(chart_files) == 3
        chart_payloads = [archive.read(name).decode("utf-8") for name in chart_files]
        assert any("<strRef>" in payload for payload in chart_payloads)
        assert any('<plotVisOnly val="0"' in payload for payload in chart_payloads[1:])


def test_api_export_results_recomputes_tampered_summary_from_form_state() -> None:
    payload = {
        "entries": [
            {
                "signature": "tampered-signature",
                "saved_at": "2026-04-18T12:00:00",
                "form_state": {
                    "layers": [
                        {"material_id": "RC416T", "theta_deg": 45.0},
                        {"material_id": "UD", "theta_deg": 0.0},
                        {"material_id": "RC416T", "theta_deg": 90.0},
                    ],
                    "is_symmetric": False,
                    "core_material_id": "Honeycomb",
                    "insert_dummy_layer_for_odd_compatibility": True,
                    "compatibility_mode": "legacy",
                    "custom_materials": [],
                    "three_point_bending": {
                        "elastic_gradient": 2649.0,
                        "rigidez_rig": 14871.0,
                        "span_m": 0.4,
                        "span_mm": 400.0,
                        "width_m": 0.275,
                        "width_mm": 275.0,
                    },
                },
                "summary": {
                    "elastic_gradient_theory": 1.0,
                    "ei_theory": 2.0,
                    "fiber_thickness_mm": 3.0,
                    "total_thickness_mm": 4.0,
                    "core_material_id": "BROKEN",
                    "is_symmetric": True,
                    "visible_layers": 99,
                },
                "result_data": {"equivalent_properties": {"e11_gpa": -1}},
            }
        ]
    }

    response = client.post("/api/export-results", json=payload)
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["3 capas"]
    assert worksheet["E30"].value == "Honeycomb"
    assert worksheet["G30"].value == 22.32
    assert worksheet["H30"].value == 1.16
    assert worksheet["I30"].value == 3543.258
